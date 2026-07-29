"""
Reporting Service
Handles generation of inventory and stock movement reports in CSV, Excel, and PDF formats.
"""

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from src.config import EXPORT_DIR
from src.services.inventory_service import InventoryService
from src.services.supplier_service import SupplierService
from src.services.time_sync_service import get_time_sync_service
from src.models.models import Item, StockMovement

logger = logging.getLogger(__name__)


class ReportingService:
    """Service for generating reports"""

    def __init__(self):
        self.inventory_service = InventoryService()
        self.supplier_service = SupplierService()
        self.time_sync_service = get_time_sync_service()
        self.export_dir = EXPORT_DIR
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def _default_filename(self, prefix: str, extension: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        return self.export_dir / f"{prefix}_{timestamp}.{extension}"

    def _save_csv(
        self,
        path: Path,
        headers: List[str],
        rows: List[List[str]],
        generated_at: Optional[str] = None,
    ) -> Tuple[bool, str]:
        try:
            with path.open(mode="w", newline="", encoding="utf-8") as fh:
                writer = csv.writer(fh)
                if generated_at:
                    writer.writerow(["Generated at", generated_at])
                    writer.writerow([])
                writer.writerow(headers)
                writer.writerows(rows)
            return True, str(path)
        except Exception as e:
            logger.error(f"Failed to save CSV report: {e}")
            return False, f"Failed to save CSV report: {e}"

    def _save_excel(
        self,
        path: Path,
        headers: List[str],
        rows: List[List[str]],
        generated_at: Optional[str] = None,
    ) -> Tuple[bool, str]:
        try:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"

            current_row = 1
            if generated_at:
                sheet.cell(row=current_row, column=1, value="Generated at").font = Font(bold=True)
                sheet.cell(row=current_row, column=2, value=generated_at)
                current_row += 2

            for column_index, header in enumerate(headers, start=1):
                cell = sheet.cell(row=current_row, column=column_index, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row_index, row in enumerate(rows, start=current_row + 1):
                for col_index, value in enumerate(row, start=1):
                    sheet.cell(row=row_index, column=col_index, value=value)

            workbook.save(path)
            return True, str(path)
        except Exception as e:
            logger.error(f"Failed to save Excel report: {e}")
            return False, f"Failed to save Excel report: {e}"

    def _save_pdf(
        self,
        path: Path,
        title: str,
        headers: List[str],
        rows: List[List[str]],
        generated_at: Optional[str] = None,
    ) -> Tuple[bool, str]:
        try:
            doc = SimpleDocTemplate(str(path), pagesize=letter)
            styles = getSampleStyleSheet()
            elements = [Paragraph(title, styles["Title"])]
            if generated_at:
                elements.extend([Spacer(1, 6), Paragraph(f"Generated at: {generated_at}", styles["Normal"])])
            elements.append(Spacer(1, 12))

            table_data = [headers] + rows
            table = Table(table_data, hAlign="LEFT")
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4B8BBE")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("ALIGN", (0, 0), (-1, -1), "LEFT"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.gray),
                    ]
                )
            )
            elements.append(table)
            doc.build(elements)
            return True, str(path)
        except Exception as e:
            logger.error(f"Failed to save PDF report: {e}")
            return False, f"Failed to save PDF report: {e}"

    def _format_item_row(self, item: Item) -> List[str]:
        supplier_name = ""
        if item.supplier_id:
            supplier = self.supplier_service.get_supplier_by_id(item.supplier_id)
            supplier_name = supplier.supplier_name if supplier else str(item.supplier_id)
        return [
            str(item.id or ""),
            item.item_name,
            item.product_code or "",
            item.barcode or "",
            item.category or "",
            item.manufacturer or "",
            supplier_name,
            item.supplier_product_code or "",
            item.unit_of_measurement or "",
            str(item.current_quantity),
            str(item.minimum_quantity),
            str(item.maximum_quantity),
            str(item.lead_time_days or ""),
            str(item.safety_stock_quantity or 0),
            "Active" if item.is_active else "Inactive",
            item.expiry_date.isoformat() if item.expiry_date else "",
            item.storage_location or "",
            item.temperature_requirement or "",
            "Yes" if item.is_controlled_drug else "No",
            "Yes" if item.requires_fridge else "No",
        ]

    def _format_movement_row(self, movement: StockMovement) -> List[str]:
        movement_time = ""
        if movement.movement_time:
            if hasattr(movement.movement_time, "strftime"):
                movement_time = movement.movement_time.strftime("%H:%M:%S")
            else:
                movement_time = str(movement.movement_time)

        movement_date = ""
        if movement.movement_date:
            if hasattr(movement.movement_date, "isoformat"):
                movement_date = movement.movement_date.isoformat()
            else:
                movement_date = str(movement.movement_date)

        return [
            str(movement.transaction_id or ""),
            str(movement.product_id),
            str(movement.batch_id or ""),
            str(movement.room_id or ""),
            movement.transaction_type,
            str(movement.quantity),
            str(movement.previous_quantity) if movement.previous_quantity is not None else "",
            str(movement.new_quantity) if movement.new_quantity is not None else "",
            str(movement.from_room_id or ""),
            str(movement.to_room_id or ""),
            movement_date,
            movement_time,
            movement.reason or "",
            movement.patient_area or "",
            movement.from_location or "",
            movement.to_location or "",
            movement.batch_number or "",
            movement.notes or "",
            str(movement.user_id),
        ]

    def generate_inventory_csv(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("inventory_report", "csv")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "ID",
            "Name",
            "Product Code",
            "Barcode",
            "Category",
            "Manufacturer",
            "Supplier",
            "Supplier Product Code",
            "Unit of Measurement",
            "Current Quantity",
            "Minimum Quantity",
            "Target Stock Level",
            "Lead Time in Days",
            "Safety Stock Quantity",
            "Active Status",
            "Expiry Date",
            "Storage Location",
            "Temperature Requirement",
            "Controlled Drug",
            "Requires Fridge",
        ]
        rows = [self._format_item_row(item) for item in self.inventory_service.get_all_items()]
        return self._save_csv(path, headers, rows, generated_at=signature)

    def generate_inventory_excel(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("inventory_report", "xlsx")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "ID",
            "Name",
            "Product Code",
            "Barcode",
            "Category",
            "Manufacturer",
            "Supplier",
            "Supplier Product Code",
            "Unit of Measurement",
            "Current Quantity",
            "Minimum Quantity",
            "Target Stock Level",
            "Lead Time in Days",
            "Safety Stock Quantity",
            "Active Status",
            "Expiry Date",
            "Storage Location",
            "Temperature Requirement",
            "Controlled Drug",
            "Requires Fridge",
        ]
        rows = [self._format_item_row(item) for item in self.inventory_service.get_all_items()]
        return self._save_excel(path, headers, rows, generated_at=signature)

    def generate_inventory_pdf(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("inventory_report", "pdf")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "ID",
            "Name",
            "Product Code",
            "Barcode",
            "Category",
            "Manufacturer",
            "Supplier",
            "Supplier Product Code",
            "Unit",
            "Qty",
            "Min",
            "Target",
            "Lead Time",
            "Safety Stock",
            "Status",
            "Expiry",
            "Location",
            "Temp",
            "Controlled",
            "Fridge",
        ]
        rows = [self._format_item_row(item) for item in self.inventory_service.get_all_items()]
        return self._save_pdf(path, "Inventory Report", headers, rows, generated_at=signature)

    def generate_movements_csv(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("stock_movements_report", "csv")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "Transaction ID",
            "Product ID",
            "Batch ID",
            "Room ID",
            "Transaction Type",
            "Quantity",
            "Previous Quantity",
            "New Quantity",
            "From Room",
            "To Room",
            "Date",
            "Time",
            "Reason",
            "Area",
            "From Location",
            "To Location",
            "Batch Number",
            "Notes",
            "User ID",
        ]
        rows = [self._format_movement_row(mov) for mov in self.inventory_service.get_stock_movements(limit=1000)]
        return self._save_csv(path, headers, rows, generated_at=signature)

    def generate_movements_excel(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("stock_movements_report", "xlsx")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "Transaction ID",
            "Product ID",
            "Batch ID",
            "Room ID",
            "Transaction Type",
            "Quantity",
            "Previous Quantity",
            "New Quantity",
            "From Room",
            "To Room",
            "Date",
            "Time",
            "Reason",
            "Area",
            "From Location",
            "To Location",
            "Batch Number",
            "Notes",
            "User ID",
        ]
        rows = [self._format_movement_row(mov) for mov in self.inventory_service.get_stock_movements(limit=1000)]
        return self._save_excel(path, headers, rows, generated_at=signature)

    def generate_movements_pdf(self, path: Optional[Path] = None) -> Tuple[bool, str]:
        path = path or self._default_filename("stock_movements_report", "pdf")
        signature = self.time_sync_service.get_signature_stamp()
        headers = [
            "Transaction ID",
            "Product ID",
            "Batch ID",
            "Room ID",
            "Type",
            "Qty",
            "Prev Qty",
            "New Qty",
            "From Room",
            "To Room",
            "Date",
            "Time",
            "Reason",
            "Area",
            "From",
            "To",
            "Batch",
            "Notes",
            "User ID",
        ]
        rows = [self._format_movement_row(mov) for mov in self.inventory_service.get_stock_movements(limit=1000)]
        return self._save_pdf(path, "Stock Movements Report", headers, rows, generated_at=signature)
